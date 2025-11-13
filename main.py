import os
from fastapi import FastAPI, HTTPException
from google.oauth2 import service_account
from googleapiclient.discovery import build
from io import BytesIO
import openpyxl

app = FastAPI()

# -------------------------------------
# Environment variables
# -------------------------------------
SERVICE_ACCOUNT_JSON = os.environ.get("SERVICE_ACCOUNT_JSON")
DRIVE_FILE_ID = os.environ.get("DRIVE_FILE_ID")  # Excel file in Drive

# -------------------------------------
# Build Google Drive client
# -------------------------------------
def get_drive_client():
    try:
        creds = service_account.Credentials.from_service_account_info(
            eval(SERVICE_ACCOUNT_JSON),
            scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        raise Exception("Google auth failed: " + str(e))

# -------------------------------------
# Read Excel sheet
# -------------------------------------
def read_excel_sheet(sheet_name: str):
    drive = get_drive_client()

    # Download Excel file from Drive
    request = drive.files().get_media(fileId=DRIVE_FILE_ID)
    file_bytes = request.execute()

    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found.")

    sheet = workbook[sheet_name]

    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))

    return rows

# -------------------------------------
# API endpoints
# -------------------------------------
@app.get("/debug")
def debug():
    """Check configuration and file access"""
    try:
        drive = get_drive_client()
        file_info = drive.files().get(fileId=DRIVE_FILE_ID, fields="id, name, mimeType").execute()
        return {
            "file_id": DRIVE_FILE_ID,
            "file_name": file_info.get("name"),
            "mime_type": file_info.get("mimeType"),
            "service_account_configured": bool(SERVICE_ACCOUNT_JSON)
        }
    except Exception as e:
        return {"error": str(e), "file_id": DRIVE_FILE_ID, "configured": bool(SERVICE_ACCOUNT_JSON)}

@app.get("/excel/sheets")
def list_sheets():
    """List all available sheet names in the Excel file"""
    try:
        drive = get_drive_client()
        request = drive.files().get_media(fileId=DRIVE_FILE_ID)
        file_bytes = request.execute()
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        return {"sheets": workbook.sheetnames}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/excel/read-sheet")
def read_sheet(sheet: str):
    """
    Reads a specific sheet from an Excel file stored in Google Drive.
    Example: sheet=units
    """
    try:
        values = read_excel_sheet(sheet)
        return {"values": values}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
